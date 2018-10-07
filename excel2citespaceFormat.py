from openpyxl import load_workbook
from openpyxl.compat import range

wb = load_workbook('C:\\Users\\vino\\Documents\\Tencent Files'
                   '\\674912850\\FileRecv\\情报学期刊&会议'
                   '\\2018年\\会议.xlsx')
ws1 = wb.active
f = open('C:\\Users\\vino\\Downloads\\cnki\\download_hy2018.txt', 'w', encoding='utf-8')

for i in range(2, ws1.max_row + 1):
    for j in range(1, ws1.max_column + 1):
        f.write(ws1.cell(1, j).value + ' ' + str(ws1.cell(i, j).value) + '\n')

    f.write('\n')
f.close() 