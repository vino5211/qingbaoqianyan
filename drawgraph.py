from openpyxl import load_workbook
from openpyxl.compat import range
import matplotlib.pyplot as plt
import networkx as nx
import os
wb = load_workbook('D:\\OneDrive\\BC论文\\18.09.12\\newbc18.9.12.xlsx')
ws1 = wb.active

file_name = {0:'rbc.png',1:'ca.png',2:'cs.png',3:'sim.png'}
for vertice in range(1,42):
    dirs = 'D:\\OneDrive\\BC论文\\18.09.12\\' + str(vertice)

    if not os.path.exists(dirs):
        os.makedirs(dirs)
    for j in range(4):
        G = nx.Graph()
        f ='D:\\OneDrive\\BC论文\\18.09.12\\'+str(vertice)+'\\'+file_name[j]

        evertex = list()
        eother = list()
        for i in range(2, ws1.max_row + 1):
            G.add_edge(ws1.cell(i, 1).value,ws1.cell(i, 2).value, weight=ws1.cell(i, j+3).value)
            if ws1.cell(i, 1).value == vertice or ws1.cell(i, 2).value == vertice:
                evertex.append((ws1.cell(i, 1).value,ws1.cell(i, 2).value))

            else:
                eother.append((ws1.cell(i, 1).value, ws1.cell(i, 2).value))
        pos = nx.circular_layout(G)  # positions for all nodes
        d = nx.degree(G)
        # nodes
        plt.figure(1, figsize=(16, 12))
        nx.draw_networkx_nodes(G, pos, node_size=[v * 100 for (k,v) in d],node_color='y')

        # edges

        oweights = [G[u][v]['weight']*50 for u, v in eother]
        nx.draw_networkx_edges(G, pos, edgelist=eother,
                                alpha=0.4, edge_color='b', style='dashdot',width=oweights,label=[str(n) for n in oweights])
        vweights = [G[u][v]['weight'] * 50 for u, v in evertex]
        nx.draw_networkx_edges(G, pos, edgelist=evertex, width=vweights,label=[str(n) for n in vweights])

        # labels
        nx.draw_networkx_labels(G, pos, font_size=20, font_family='sans-serif')
        vedgelabel={}
        for u,v in evertex:
            vedgelabel[(u,v)]=G[u][v]['weight']
        nx.draw_networkx_edge_labels(G, pos,edge_labels=vedgelabel)

        plt.axis('off')
        # plt.show()
        plt.savefig(f)
        plt.clf()
        G.clear()
