#提取基金号
from openpyxl import load_workbook
from openpyxl.compat import range
import regex as re

wb = load_workbook('C:\\Users\\A\\Documents\\Tencent Files'
                   '\\674912850\\FileRecv'
                   '\\2017.xlsx')
ws = wb.active

for i in range(2, ws.max_row + 1):
    if ws.cell(i, 10).value is not None:
        pattern = re.compile(r"[\\(（](.*?)[\\)）]")#[\\(（](.*?)[\\)）]
        result=pattern.findall(ws.cell(i,10).value)
        ws.cell(i,16).value=str(result)
wb.save('C:\\Users\\A\\Documents\\Tencent Files'
                   '\\674912850\\FileRecv'
                   '\\2017_fund.xlsx')
