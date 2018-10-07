from openpyxl import load_workbook
from openpyxl.compat import range
import os

wb = load_workbook('D:\\OneDrive\\BC论文\\18.09.12\\newbc18.9.12.xlsx')
ws1 = wb.active
f1 = open('D:\\OneDrive\\BC论文\\18.09.12\\template.net', 'r', encoding='utf-8')
s=f1.read()
f1.close()
file_name = {0:'rbc.net',1:'ca.net',2:'cs.net',3:'sim.net'}
for vertice in range(1,42):

    os.makedirs('D:\\OneDrive\\BC论文\\18.09.12\\'+str(vertice))
    for j in range(4):
        f = open('D:\\OneDrive\\BC论文\\18.09.12\\'+str(vertice)+'\\'+file_name[j], 'w', encoding='utf-8')
        f.write(s)

        for i in range(2, ws1.max_row + 1):
            if ws1.cell(i, j*7+1).value == vertice or ws1.cell(i, j*7+2).value == vertice:
                f.write(str(ws1.cell(i, j*7+1).value)+' '+str(ws1.cell(i, j*7+2).value)+' '+
                        str(ws1.cell(i, j*7+3).value)+' p Solid c Blue\n')

            else:
                f.write(str(ws1.cell(i, j*7+1).value)+' '+str(ws1.cell(i, j*7+2).value)+' '+
                        str(ws1.cell(i, j*7+3).value)+ ' p Dots c Gray\n')
        f.close()